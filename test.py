import sys
import os
from openpyxl import Workbook
import pdfplumber
from operator import itemgetter
from openpyxl.styles import Font

def parse_pdf(x_tolerance,y_tolerance,path =None,savePath = None,data = None):
    '''
    function : 处理pdf
    :param:词间最大间距，行间最大间距，输入路径，输出路径
    :return  无
    '''
    # 读入文件
    if path != None:
        pdf = pdfplumber.open(path)
    elif data != None:
        pdf = pdfplumber.load(data)
    # 记录page行数
    pdfRowNumber = 0

    wb = Workbook()
    ws = wb.active
    
    for page in pdf.pages:
        pageContainer = [] #存储所有该page的字典
        theMaxColNum = 0 #记录最大列数
        words = page.extract_words(x_tolerance =x_tolerance ,y_tolerance =y_tolerance ,keep_blank_chars =True)
        pageContainer ,theMaxColNum = compileByRowLocation(words ,x_tolerance ,y_tolerance)
        # 按照位置信息排序
        for line in pageContainer:
            line.sort(key =itemgetter('x0'))
        pageContainer.sort(key =lambda line : line[0]['top'])
        # 对前排进行判断是否出现有表头或多行为一行的情况
        if len(pageContainer[0]) < theMaxColNum:
            for i in range(len(pageContainer)):
                if len(pageContainer[i]) == theMaxColNum:
                    repairList,repairNum = align_front_row(pageContainer[0:i+1],theMaxColNum)
                    for i in range(repairNum):
                        del pageContainer[0]
                    pageContainer.insert(0,repairList)
                    break
        # 对最后一排进行判断
        if len(pageContainer[-1]) < theMaxColNum:
            pageContainer[-1] = align_last_row(pageContainer[-2:],theMaxColNum)
        # 写入excel
        # ftTitle = Font(name='font',size=14)
        # ftText = Font(name = '',size=10)
        for idx,line in enumerate(pageContainer):
            for idy,item in enumerate(line):
                cellIndex = ws.cell(row = idx + 1 + pdfRowNumber, column = idy + 1)
                if item['text'] == '':
                    pass
                elif item['text'] == None:
                    ws.merge_cells(start_row=idx + 1 + pdfRowNumber, start_column=1, end_row=idx + 1 + pdfRowNumber, end_column=len(line))
                    break
                else:
                    # cellIndex.font = ftText
                    if item['text'].isdigit():
                        cellIndex.value = int(item['text'])
                    elif is_float(item['text']):
                        cellIndex.value = float(item['text'])
                    else:
                        cellIndex.value = item['text']
        # 将该页的行数相加，使excel连续
        pdfRowNumber += len(pageContainer)
    # 保存excel文件至本地
    if savePath != None:
        wb.save(savePath)
    else:
        wb.save(path.replace('.pdf','.xlsx'))

def align_front_row(rowList ,theMaxColNum):
    '''
    function : 对输入的数行进行整理，判断是否有表头或多行处于一行的情况
    :param:多行数据的列表，最大列数
    :return  处理后数据列表，具有最大列数的行数
    '''
    locForMax = []
    theMaxRowLoc = 0
    adjustForMin = []
    # 初始化返回数组
    for i in range(theMaxColNum):
        adjustForMin.append({'text':''})
    # 得到最大列数的位置信息
    for idx,row in enumerate(rowList):
        if len(row) == theMaxColNum:
            theMaxRowLoc = idx
            for item in row:
                locForMax.append([item['x0'] ,item['x1']])
            break
    # 判断表头，如无表头，则判断并将多行合并成一行
    if theMaxRowLoc == 1 and len(rowList[0]) == 1:
        for i in range(theMaxColNum - 1):
            rowList[0].append({'text':None})
        return rowList[0],theMaxRowLoc
    else:
        for row in rowList[0:theMaxRowLoc]:
            for item in row:
                for idx,itemInMax in enumerate(locForMax):
                    if item['x0'] > itemInMax[1] or itemInMax[0] > item['x1']:
                        pass
                    else:
                        adjustForMin[idx]['text'] += item['text']
        return adjustForMin,theMaxRowLoc

def align_last_row(rowList,theMaxColNum):
    '''
    function : 判断最后是否有单独的数据
    :param:最后两行数据，最大列数
    :return  修改后最后一行数据
    '''
    adjustForMin = []

    for i in range(theMaxColNum):
        adjustForMin.append({'text':''})
    
    for item in rowList[1]:
        for idx,itemInMax in enumerate(rowList[0]):
            if item['x0'] > itemInMax['x1'] or itemInMax['x0'] > item['x1']:
                pass
            else:
                adjustForMin[idx]['text'] = item['text']
    return adjustForMin

def compileByRowLocation(rowList ,x_tolerance ,y_tolerance):
    '''
    function : 将得到的字典按照行排序
    :param:具有所有字典信息的列表，词间最大间距，行间最大间距
    :return  处理后二维list结果，最大列数
    '''
    pageContainer = [] #存储所有该page的字典
    pageRowsLocation = [] #存储每一行的位置（top）
    theMaxColNum = 0 #记录最大列数
    # 将字典添加入pageContainer
    countColNum = 0
    for item in rowList:
        flagExist = is_in(pageRowsLocation ,item['top'] ,y_tolerance)
        if flagExist == None:
            pageRowsLocation.append(item['top'])
            pageContainer.append([item])
            
            if countColNum > theMaxColNum:
                theMaxColNum = countColNum
            countColNum = 1
        else:
            pageContainer[flagExist].append(item)
            countColNum += 1
    
    return pageContainer,theMaxColNum
    
def is_in(rowList ,num ,y_tolerance):
    '''
    function : 判断该字典已经有对应的行在pagecontainer中
    :param:输入当前行位置信息（pageRowsLocation），行位置信息（top），行间最大间距
    :return  如果已经有该行，返回该行位置，如没有返回None
    '''
    for idx ,item in enumerate(rowList):
        if abs(num -item) < y_tolerance:
            return idx
    return None

def is_float(str):
    '''
    function : 判断是否浮点数字
    :param:字符串
    :return  bool值
    '''
    try:
        float(str)
        return True
    except ValueError:
        return False

def getfiles(ospath):
    try:
        files = os.listdir(ospath)
        for f in files:
            mypath = os.path.join(ospath, f)
            # print(mypath)
            if os.path.isfile(mypath):
                ext = os.path.splitext(mypath)
                if ext[1] == '.pdf' :#指定文件类型
                    parse_pdf(1,1.5,path= mypath) 
                    print('处理完' + mypath)                    
            if os.path.isdir(mypath):
                # print('heihei'+mypath)
                getfiles(mypath)#递归
    except  Exception as e :
        print(str(e))

if __name__ == '__main__':
    # savePath = r'C:\Users\15644\Desktop\test_res_plumber\test_3.xlsx'
    path = r'C:\Users\15644\Desktop\pdf_file\test_pdf_list\test_1.pdf'
    f = open(path,'rb')
    # # print(f.read())
    parse_pdf(1,1.5,data=f,savePath=r"C:\Users\15644\Desktop\test_1.xlsx")
    # getfiles(r'C:\Users\15644\Desktop\6月大波浪')