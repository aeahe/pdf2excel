from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfdevice import PDFDevice
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine, LTFigure, LTImage, LTChar, LTText,LTPage
import sys
import os
from openpyxl import Workbook
from operator import itemgetter
from openpyxl.styles import Alignment

def parse_pdf(path = None,data = None,savePath = None,y_tolerance = 1.5,char_tolerance = 0.5):
    '''
    function : 处理pdf
    :param:词间最大间距，行间最大间距，输入路径，输出路径
    :return  无
    '''   
    # 记录page行数
    pdfRowNumber = 0

    theMaxColSize = []

    wb = Workbook()
    ws = wb.active

    if data == None:
        data = open(path, 'rb')
    
    parser = PDFParser(data)
    document = PDFDocument(parser)
    rsrcmgr = PDFResourceManager()
    device = PDFPageAggregator(rsrcmgr, laparams= None)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    for page in PDFPage.create_pages(document):
        interpreter.process_page(page)
        height = page.mediabox[3]-page.mediabox[1]
        layout = device.get_result()
        pageContainer,theMaxColNum = get_line_word(layout,height,y_tolerance = y_tolerance,char_tolerance= char_tolerance)
        # 按照位置信息排序
        for line in pageContainer:
            line.sort(key =itemgetter('x0'))
        pageContainer.sort(key =lambda line : line[0]['top'])

        if len(pageContainer[0]) < theMaxColNum:
            for i in range(len(pageContainer)):
                if len(pageContainer[i]) == theMaxColNum:
                    repairList = align_front_row(pageContainer[0:i],theMaxColNum)
                    del pageContainer[0:i]
                    pageContainer.insert(0,repairList)
                    break
        # 对最后一排进行判断
        if len(pageContainer[-1]) < theMaxColNum:
            pageContainer[-1] = align_last_row(pageContainer[-2:],theMaxColNum)
        # 写入excel
        alignment = Alignment(horizontal='center',vertical= 'center')
        for idx,line in enumerate(pageContainer):
            for idy,item in enumerate(line):
                cellIndex = ws.cell(row = idx + 1 + pdfRowNumber, column = idy + 1)
                if item['text'] == '':
                    pass
                elif item['text'] == None:
                    ws.merge_cells(start_row=idx + 1 + pdfRowNumber, start_column=1, end_row=idx + 1 + pdfRowNumber, end_column=theMaxColNum)
                    ws.cell(idx + 1 + pdfRowNumber,1).alignment = alignment
                    break
                else:
                    if idx == 0 and len(line) == 2:
                        pass
                    else:
                        cellIndex.alignment = alignment                    

                    if item['text'].isdigit():
                        cellIndex.value = int(item['text'])
                        cellIndex.number_format = '0'
                    elif is_float(item['text']):
                        cellIndex.value = float(item['text'])
                    else:
                        cellIndex.value = item['text']

        thePageMaxColSize = [0 for i in range(theMaxColNum)]
        for line in pageContainer:
            if len(line) == 2:
                continue
            for col,item in enumerate(line):
                if len(item['text']) > thePageMaxColSize[col]:
                    thePageMaxColSize[col] = len(item['text'])

        if theMaxColSize == []:
            theMaxColSize = thePageMaxColSize[:]
        else:
            for i in range(theMaxColNum):
                if theMaxColSize[i] < thePageMaxColSize[i]:
                    theMaxColSize[i] = thePageMaxColSize[i]
        # 将该页的行数相加，使excel连续
        pdfRowNumber += len(pageContainer)

    # 保存excel文件至本地
    letter = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    for col,theSize in enumerate(theMaxColSize):
        rest = (col + 1) % 26
        cut = int((col + 1) / 26)
        colLetter = ''
        if cut == 0:
            colLetter = letter[rest - 1]
        else:
            colLetter = letter[cut] + letter[rest - 1]
        ws.column_dimensions[colLetter].width = theSize*2

    if savePath != None:
        wb.save(savePath)
    else:
        wb.save(path.replace('.pdf','.xlsx'))

def align_front_row(rowList ,theMaxColNum):
    '''
    function : 对输入的数行进行整理，判断是否有表头或多行处于一行的情况
    :param:多行数据的列表，最大列数
    :return  处理后数据列表，具有最大列数的行数
    '''
    def insert_into_res(item,locs,indexForLocs):
        '''
        function : 将词插入合适位置
        :param:词，位置信息，装有词内容的列表
        :return  更新好的位置信息和装有词内容的列表
        '''
        if locs == []:
            locs.append([item['x0'],item['x1']])
            indexForLocs.append(item)
        else:
            flag = False
            for idx,loc in enumerate(locs):
                if (item['x0'] > loc[1] or loc[0] > item['x1']) == False:
                    indexForLocs[idx] = wordBox2word([indexForLocs[idx],item])
                    locs[idx] = [indexForLocs[idx]['x0'],indexForLocs[idx]['x1']]
                    flag = True
                    break
            if flag == False:
                locs.append([item['x0'],item['x1']])
                indexForLocs.append(item)
          
        return locs,indexForLocs

    # 判断表头，如无表头，则判断并将多行合并成一行
    if len(rowList) == 1 and len(rowList[0]) == 1:
        rowList[0].append({'text':None})
        return rowList[0]
    else:
        locs = []
        indexForLocs = []
        for line in rowList:
            for item in line:
                locs,indexForLocs = insert_into_res(item,locs,indexForLocs)
        indexForLocs.sort(key = itemgetter('x0'))
    
    return indexForLocs

def align_last_row(rowList,theMaxColNum):
    '''
    function : 判断最后是否有单独的数据
    :param:最后两行数据，最大列数
    :return  修改后最后一行数据
    '''
    adjustForMin = [{'text':''} for i in range(theMaxColNum)]
    
    for item in rowList[1]:
        for idx,itemInMax in enumerate(rowList[0]):
            if item['x0'] > itemInMax['x1'] or itemInMax['x0'] > item['x1']:
                pass
            else:
                adjustForMin[idx]['text'] = item['text']
    return adjustForMin

def is_float(str):
    '''
    function : 判断是否浮点数字
    :param:字符串
    :return  bool值
    '''
    try:
        float(str)
        return True
    except ValueError:
        return False

def getfiles(ospath):
    try:
        files = os.listdir(ospath)
        for f in files:
            mypath = os.path.join(ospath, f)
            # print(mypath)
            if os.path.isfile(mypath):
                ext = os.path.splitext(mypath)
                if ext[1] == '.pdf' :#指定文件类型
                    parse_pdf(path= mypath) 
                    print('处理完' + mypath)                    
            if os.path.isdir(mypath):
                # print('heihei'+mypath)
                getfiles(mypath)#递归
    except  Exception as e :
        print(str(e))

def wordBox2word(wordBox):
    '''
    function : 将字符组成词
    :param:存有字符的列表
    :return  词
    '''
    word = {
        'x0' :min(map(itemgetter('x0'), wordBox)),
        'x1' : max(map(itemgetter("x1"), wordBox)),
        "top": min(map(itemgetter("top"), wordBox)),
        "bottom": max(map(itemgetter("bottom"), wordBox)),
        'text': ''.join(map(itemgetter("text"), wordBox)),
        'size': wordBox[0]['size'],
        'font': wordBox[0]['font'],
    }
    return word

def get_line_word(ltChars,height,y_tolerance = 1.5,char_tolerance = 0.5):
    '''
    function : 将字符组成词，词组成行
    :param:所有字符，page高度，词间距
    :return  处理后page内容
    '''
    lines = []
    line = []
    wordBox = []
    theMaxColNum = 0
    lastCharX1 = 0
    lastCharTop = 0
    lastPos = len(ltChars) - 1
    
    # ltChars = ltChars.get_chars()
    # ltChars.sort(key = lambda x: (height - x.y1 ,x.x0))

    for idx,ltChar in enumerate(ltChars):
        if isinstance(ltChar,LTChar):
            x0 =ltChar.x0
            x1= ltChar.x1
            top = height - ltChar.y1
            changeLtChar = {
                    'x0':x0,
                    'x1':x1,
                    'top':top,
                    'bottom':height - ltChar.y0,
                    'font':ltChar.fontname,
                    'size':ltChar.size,
                    'text':ltChar.get_text()
                }
            if wordBox == []:
                lastCharX1 = x1
                lastCharTop = top
                wordBox.append(changeLtChar)
            else:
                if 0 <= abs(x0 - lastCharX1) < char_tolerance:
                    lastCharX1 = x1
                    lastCharTop = top
                    wordBox.append(changeLtChar)
                else:
                    line.append(wordBox2word(wordBox))
                    wordBox = []
                    wordBox.append(changeLtChar)
                
                    if y_tolerance < abs(top - lastCharTop):
                        if len(line) > theMaxColNum:
                            theMaxColNum = len(line)
                        lines.append(line)
                        line = []

                    lastCharX1 = x1
                    lastCharTop = top
                
                if idx == lastPos:
                    line.append(wordBox2word(wordBox))
                    if len(line) > theMaxColNum:
                            theMaxColNum = len(line)
                    lines.append(line)

    return lines,theMaxColNum
        
if __name__ == '__main__':
    savePath = r'C:\Users\15644\Desktop\test_12.xlsx'
    path = r'C:\Users\15644\Desktop\pdf_file\test_pdf_list\test_12.pdf'
    parse_pdf(path=path,savePath=savePath)